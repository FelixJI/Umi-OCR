# ===============================================
# =============== 表格格式转换器 ===============
# ===============================================

"""
表格输出格式转换器
支持将表格数据转换为多种格式：
- Markdown
- JSON
- HTML
- Excel (xlsx)
"""

import json
import logging
from typing import Dict, List, Any, Union, Optional
from pathlib import Path

logger = logging.getLogger(__name__)


class TableConverter:
    """
    表格输出格式转换器
    
    将表格数据转换为多种输出格式，支持复杂表格结构。
    """
    
    @staticmethod
    def to_markdown(table_data: Union[Dict, List], title: str = "") -> str:
        """
        将表格数据转换为Markdown格式
        
        Args:
            table_data: 表格数据（可以是单个表格或表格列表）
            title: 可选的表格标题
            
        Returns:
            Markdown格式的表格字符串
        """
        if isinstance(table_data, list):
            # 多个表格
            md_parts = []
            for i, table in enumerate(table_data):
                table_title = title or f"表格 {i + 1}"
                md_parts.append(f"### {table_title}\n")
                md_parts.append(TableConverter._single_table_to_markdown(table))
            return "\n\n".join(md_parts)
        else:
            # 单个表格
            if title:
                return f"### {title}\n\n" + TableConverter._single_table_to_markdown(table_data)
            return TableConverter._single_table_to_markdown(table_data)
    
    @staticmethod
    def _single_table_to_markdown(table: Dict) -> str:
        """将单个表格转换为Markdown"""
        # 支持多种输入格式
        
        # 格式1: HTML表格
        if "html" in table:
            return TableConverter._html_to_markdown(table["html"])
        
        # 格式2: cells列表
        if "cells" in table:
            return TableConverter._cells_to_markdown(table["cells"])
        
        # 格式3: rows列表（二维数组）
        if "rows" in table:
            return TableConverter._rows_to_markdown(table["rows"])
        
        # 格式4: 直接是二维列表
        if isinstance(table, list) and len(table) > 0 and isinstance(table[0], list):
            return TableConverter._rows_to_markdown(table)
        
        return ""
    
    @staticmethod
    def _html_to_markdown(html: str) -> str:
        """将HTML表格转换为Markdown"""
        try:
            from html.parser import HTMLParser
            
            class TableParser(HTMLParser):
                def __init__(self):
                    super().__init__()
                    self.rows = []
                    self.current_row = []
                    self.current_cell = ""
                    self.in_cell = False
                    self.is_header = False
                
                def handle_starttag(self, tag, attrs):
                    if tag == "th":
                        self.in_cell = True
                        self.is_header = True
                        self.current_cell = ""
                    elif tag == "td":
                        self.in_cell = True
                        self.current_cell = ""
                    elif tag == "tr":
                        self.current_row = []
                
                def handle_endtag(self, tag):
                    if tag in ("td", "th"):
                        self.in_cell = False
                        cell_text = self.current_cell.strip().replace("|", "\\|")
                        cell_text = cell_text.replace("\n", " ")
                        self.current_row.append(cell_text)
                    elif tag == "tr":
                        if self.current_row:
                            self.rows.append(self.current_row)
                
                def handle_data(self, data):
                    if self.in_cell:
                        self.current_cell += data
            
            parser = TableParser()
            parser.feed(html)
            
            if not parser.rows:
                return ""
            
            return TableConverter._rows_to_markdown(parser.rows)
            
        except Exception as e:
            logger.warning(f"HTML表格转Markdown失败: {e}")
            return ""
    
    @staticmethod
    def _cells_to_markdown(cells: List[Dict]) -> str:
        """将cells数据转换为Markdown"""
        if not cells:
            return ""
        
        # 计算表格尺寸
        max_row = max(c.get("row", 0) for c in cells) + 1
        max_col = max(c.get("col", 0) for c in cells) + 1
        
        # 创建表格矩阵
        table = [["" for _ in range(max_col)] for _ in range(max_row)]
        
        for cell in cells:
            row = cell.get("row", 0)
            col = cell.get("col", 0)
            text = str(cell.get("text", "")).replace("|", "\\|").replace("\n", " ")
            
            # 处理合并单元格
            row_span = cell.get("row_span", 1)
            col_span = cell.get("col_span", 1)
            
            if 0 <= row < max_row and 0 <= col < max_col:
                table[row][col] = text
        
        return TableConverter._rows_to_markdown(table)
    
    @staticmethod
    def _rows_to_markdown(rows: List[List[str]]) -> str:
        """将行数据转换为Markdown表格"""
        if not rows:
            return ""
        
        # 计算最大列数
        max_cols = max(len(row) for row in rows)
        
        # 规范化所有行
        normalized_rows = []
        for row in rows:
            normalized_row = list(row) + [""] * (max_cols - len(row))
            normalized_rows.append(normalized_row)
        
        md_lines = []
        
        # 表头
        header = normalized_rows[0]
        md_lines.append("| " + " | ".join(str(cell) for cell in header) + " |")
        md_lines.append("| " + " | ".join(["---"] * max_cols) + " |")
        
        # 表体
        for row in normalized_rows[1:]:
            md_lines.append("| " + " | ".join(str(cell) for cell in row) + " |")
        
        return "\n".join(md_lines)
    
    @staticmethod
    def to_json(table_data: Union[Dict, List], indent: int = 2) -> str:
        """
        将表格数据转换为JSON格式
        
        Args:
            table_data: 表格数据
            indent: JSON缩进空格数
            
        Returns:
            JSON格式字符串
        """
        # 标准化表格数据
        if isinstance(table_data, list):
            standardized = [TableConverter._standardize_table(t) for t in table_data]
        else:
            standardized = TableConverter._standardize_table(table_data)
        
        return json.dumps(standardized, ensure_ascii=False, indent=indent)
    
    @staticmethod
    def _standardize_table(table: Dict) -> Dict:
        """标准化表格数据结构"""
        result = {
            "rows": [],
            "metadata": {},
        }
        
        # 提取行数据
        if "rows" in table:
            result["rows"] = table["rows"]
        elif "cells" in table:
            result["rows"] = TableConverter._cells_to_rows(table["cells"])
        elif "html" in table:
            result["rows"] = TableConverter._html_to_rows(table["html"])
        elif isinstance(table, list):
            result["rows"] = table
        
        # 提取元数据
        for key in ["bbox", "confidence", "type"]:
            if key in table:
                result["metadata"][key] = table[key]
        
        return result
    
    @staticmethod
    def _cells_to_rows(cells: List[Dict]) -> List[List[str]]:
        """将cells数据转换为行列表"""
        if not cells:
            return []
        
        max_row = max(c.get("row", 0) for c in cells) + 1
        max_col = max(c.get("col", 0) for c in cells) + 1
        
        table = [["" for _ in range(max_col)] for _ in range(max_row)]
        
        for cell in cells:
            row = cell.get("row", 0)
            col = cell.get("col", 0)
            text = str(cell.get("text", ""))
            if 0 <= row < max_row and 0 <= col < max_col:
                table[row][col] = text
        
        return table
    
    @staticmethod
    def _html_to_rows(html: str) -> List[List[str]]:
        """将HTML表格转换为行列表"""
        try:
            from html.parser import HTMLParser
            
            class TableParser(HTMLParser):
                def __init__(self):
                    super().__init__()
                    self.rows = []
                    self.current_row = []
                    self.current_cell = ""
                    self.in_cell = False
                
                def handle_starttag(self, tag, attrs):
                    if tag in ("td", "th"):
                        self.in_cell = True
                        self.current_cell = ""
                    elif tag == "tr":
                        self.current_row = []
                
                def handle_endtag(self, tag):
                    if tag in ("td", "th"):
                        self.in_cell = False
                        self.current_row.append(self.current_cell.strip())
                    elif tag == "tr":
                        if self.current_row:
                            self.rows.append(self.current_row)
                
                def handle_data(self, data):
                    if self.in_cell:
                        self.current_cell += data
            
            parser = TableParser()
            parser.feed(html)
            return parser.rows
            
        except Exception:
            return []
    
    @staticmethod
    def to_html(table_data: Union[Dict, List], styled: bool = True) -> str:
        """
        将表格数据转换为HTML格式
        
        Args:
            table_data: 表格数据
            styled: 是否包含内联样式
            
        Returns:
            HTML格式字符串
        """
        style = ""
        if styled:
            style = """
<style>
    .table-container { margin: 20px 0; }
    .table-container table { 
        border-collapse: collapse; 
        width: 100%; 
        font-family: Arial, sans-serif;
    }
    .table-container th, .table-container td { 
        border: 1px solid #ddd; 
        padding: 8px; 
        text-align: left; 
    }
    .table-container th { 
        background-color: #4CAF50; 
        color: white; 
    }
    .table-container tr:nth-child(even) { 
        background-color: #f2f2f2; 
    }
    .table-container tr:hover { 
        background-color: #ddd; 
    }
</style>
"""
        
        html_parts = [style]
        
        if isinstance(table_data, list):
            for i, table in enumerate(table_data):
                html_parts.append(f'<div class="table-container">')
                html_parts.append(f'<h3>表格 {i + 1}</h3>')
                html_parts.append(TableConverter._single_table_to_html(table))
                html_parts.append('</div>')
        else:
            html_parts.append('<div class="table-container">')
            html_parts.append(TableConverter._single_table_to_html(table_data))
            html_parts.append('</div>')
        
        return "\n".join(html_parts)
    
    @staticmethod
    def _single_table_to_html(table: Dict) -> str:
        """将单个表格转换为HTML"""
        # 如果已经是HTML
        if "html" in table:
            return table["html"]
        
        # 转换为行数据
        if "rows" in table:
            rows = table["rows"]
        elif "cells" in table:
            rows = TableConverter._cells_to_rows(table["cells"])
        elif isinstance(table, list):
            rows = table
        else:
            return ""
        
        if not rows:
            return ""
        
        html_lines = ["<table>"]
        
        # 表头
        html_lines.append("  <thead>")
        html_lines.append("    <tr>")
        for cell in rows[0]:
            html_lines.append(f"      <th>{cell}</th>")
        html_lines.append("    </tr>")
        html_lines.append("  </thead>")
        
        # 表体
        if len(rows) > 1:
            html_lines.append("  <tbody>")
            for row in rows[1:]:
                html_lines.append("    <tr>")
                for cell in row:
                    html_lines.append(f"      <td>{cell}</td>")
                html_lines.append("    </tr>")
            html_lines.append("  </tbody>")
        
        html_lines.append("</table>")
        
        return "\n".join(html_lines)
    
    @staticmethod
    def to_excel(
        table_data: Union[Dict, List],
        output_path: str,
        sheet_name: str = "Sheet1"
    ) -> bool:
        """
        将表格数据转换为Excel文件
        
        Args:
            table_data: 表格数据
            output_path: 输出文件路径
            sheet_name: 工作表名称
            
        Returns:
            是否成功
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            
            wb = Workbook()
            
            # 处理多个表格
            if isinstance(table_data, list):
                tables = table_data
            else:
                tables = [table_data]
            
            for i, table in enumerate(tables):
                if i == 0:
                    ws = wb.active
                    ws.title = sheet_name if len(tables) == 1 else f"{sheet_name}_{i + 1}"
                else:
                    ws = wb.create_sheet(title=f"{sheet_name}_{i + 1}")
                
                # 获取行数据
                if "rows" in table:
                    rows = table["rows"]
                elif "cells" in table:
                    rows = TableConverter._cells_to_rows(table["cells"])
                elif "html" in table:
                    rows = TableConverter._html_to_rows(table["html"])
                elif isinstance(table, list):
                    rows = table
                else:
                    continue
                
                # 写入数据
                for row_idx, row in enumerate(rows, 1):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        
                        # 表头样式
                        if row_idx == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(
                                start_color="4CAF50",
                                end_color="4CAF50",
                                fill_type="solid"
                            )
                        
                        # 边框
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.border = thin_border
                        cell.alignment = Alignment(wrap_text=True)
                
                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存文件
            wb.save(output_path)
            logger.info(f"Excel文件已保存: {output_path}")
            return True
            
        except ImportError:
            logger.error("需要安装openpyxl库: pip install openpyxl")
            return False
        except Exception as e:
            logger.error(f"导出Excel失败: {e}", exc_info=True)
            return False
    
    @staticmethod
    def convert(
        table_data: Union[Dict, List],
        output_format: str,
        output_path: Optional[str] = None,
        **kwargs
    ) -> Union[str, bool]:
        """
        统一的格式转换接口
        
        Args:
            table_data: 表格数据
            output_format: 输出格式 (markdown/json/html/excel)
            output_path: 输出文件路径（excel格式必需）
            **kwargs: 额外参数
            
        Returns:
            转换结果（字符串）或成功标志（excel）
        """
        output_format = output_format.lower()
        
        if output_format == "markdown":
            return TableConverter.to_markdown(table_data, **kwargs)
        elif output_format == "json":
            return TableConverter.to_json(table_data, **kwargs)
        elif output_format == "html":
            return TableConverter.to_html(table_data, **kwargs)
        elif output_format == "excel":
            if not output_path:
                raise ValueError("Excel格式需要指定output_path")
            return TableConverter.to_excel(table_data, output_path, **kwargs)
        else:
            raise ValueError(f"不支持的输出格式: {output_format}")
