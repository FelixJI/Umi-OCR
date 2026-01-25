# ===============================================
# =============== 发票识别 - 任务管理器 ===============
# ===============================================

"""
发票识别任务管理器
支持多种发票类型：
- 增值税发票（专票/普票/电子发票）
- 机票行程单
- 火车票
- 出租车发票
- 通用票据
"""

import os
import json
from typing import Dict, Any, List, Optional

from umi_log import logger
from .mission import Mission
from ..ocr.engines import (
    get_engine_class,
    get_cloud_engines,
    BaiduOCREngine,
    TencentOCREngine,
    AlibabaOCREngine,
)

# 合法文件后缀
ImageSuf = [
    ".jpg",
    ".jpe",
    ".jpeg",
    ".jfif",
    ".png",
    ".webp",
    ".bmp",
    ".tif",
    ".tiff",
]

DocSuf = [
    ".pdf",
]

# 发票类型定义
INVOICE_TYPES = {
    "vat_invoice": {
        "name": "增值税发票",
        "description": "增值税专用发票、普通发票、电子发票",
        "fields": [
            "发票代码", "发票号码", "开票日期", "校验码",
            "销售方名称", "销售方税号", "购买方名称", "购买方税号",
            "合计金额", "合计税额", "价税合计", "备注"
        ],
    },
    "invoice": {
        "name": "通用发票识别",
        "description": "自动识别多种类型发票",
        "fields": [],
    },
    "train_ticket": {
        "name": "火车票",
        "description": "铁路客票、高铁票",
        "fields": [
            "出发站", "到达站", "车次", "座位类型",
            "票价", "乘车日期", "姓名", "身份证号"
        ],
    },
    "taxi_invoice": {
        "name": "出租车发票",
        "description": "出租车机打发票",
        "fields": [
            "发票代码", "发票号码", "上车时间", "下车时间",
            "里程", "金额", "车牌号"
        ],
    },
    "air_ticket": {
        "name": "机票行程单",
        "description": "航空运输电子客票行程单",
        "fields": [
            "旅客姓名", "身份证号", "航班号", "出发城市", "到达城市",
            "航班日期", "票价", "燃油附加费", "民航发展基金", "合计"
        ],
    },
    "quota_invoice": {
        "name": "定额发票",
        "description": "通用定额发票",
        "fields": ["发票代码", "发票号码", "金额"],
    },
    "receipt": {
        "name": "收据/小票",
        "description": "通用收据、购物小票",
        "fields": [],
    },
}


class __MissionInvoiceClass(Mission):
    """
    发票识别任务管理器
    
    支持云端API（百度/腾讯/阿里云）进行发票识别。
    可识别增值税发票、火车票、出租车发票、机票行程单等。
    """
    
    def __init__(self):
        super().__init__()
        self._engine_type = "baidu_ocr"  # 默认使用百度OCR
        self._engine = None
        self._invoice_type = "vat_invoice"
        self._results = []
    
    # ========================= 【重载】 =========================
    
    def addMissionList(self, msnInfo, msnList):
        """添加任务列表"""
        argd = msnInfo.get("argd", {})
        self._invoice_type = argd.get("invoice.type", "vat_invoice")
        self._engine_type = argd.get("invoice.engine", "baidu_ocr")
        
        # 检查任务合法性
        for i in range(len(msnList) - 1, -1, -1):
            if "path" in msnList[i]:
                p = msnList[i]["path"]
                ext = os.path.splitext(p)[-1].lower()
                if ext not in ImageSuf and ext not in DocSuf:
                    logger.warning(f"添加发票识别任务时，第{i}项的路径不是支持的文件类型：{p}")
                    del msnList[i]
            elif "bytes" not in msnList[i] and "base64" not in msnList[i]:
                logger.warning(f"添加发票识别任务时，第{i}项不含 path、bytes、base64")
                del msnList[i]
        
        return super().addMissionList(msnInfo, msnList)
    
    def msnPreTask(self, msnInfo):
        """用于更新api和参数"""
        argd = msnInfo.get("argd", {})
        engine_type = argd.get("invoice.engine", self._engine_type)
        
        # 检查引擎是否变化
        if not self._engine or self._engine_type != engine_type:
            self._engine_type = engine_type
            self._engine = self._create_engine(argd)
            
            if not self._engine:
                return f"[Error] 发票识别引擎创建失败: {engine_type}"
            
            # 初始化引擎
            if not self._engine.initialize():
                return f"[Error] 发票识别引擎初始化失败: {engine_type}"
        
        return ""
    
    def msnTask(self, msnInfo, msn):
        """执行发票识别任务"""
        argd = msnInfo.get("argd", {})
        invoice_type = argd.get("invoice.type", self._invoice_type)
        
        try:
            # 准备识别参数
            kwargs = {"api_type": invoice_type}
            
            # 执行识别
            if "path" in msn:
                res = self._engine.recognize(msn["path"], **kwargs)
                res["path"] = msn["path"]
            elif "bytes" in msn:
                res = self._engine.recognize(msn["bytes"], **kwargs)
            elif "base64" in msn:
                import base64
                image_bytes = base64.b64decode(msn["base64"])
                res = self._engine.recognize(image_bytes, **kwargs)
            else:
                res = {
                    "code": 901,
                    "data": "[Error] Unknown task type.",
                }
            
            # 格式化发票结果
            if res.get("code") == 100:
                res = self._format_invoice_result(res, invoice_type)
            
            # 保存结果
            if res.get("code") == 100:
                self._results.append(res)
            
            return res
            
        except Exception as e:
            logger.error(f"发票识别任务执行失败: {e}", exc_info=True)
            return {
                "code": 905,
                "data": f"[Error] 发票识别失败: {e}",
            }
    
    def _create_engine(self, config: Dict[str, Any]):
        """创建识别引擎"""
        engine_class = get_engine_class(self._engine_type)
        if not engine_class:
            logger.error(f"未找到引擎类型: {self._engine_type}")
            return None
        
        # 构建引擎配置
        engine_config = {
            "api_key": config.get("invoice.api_key", ""),
            "secret_key": config.get("invoice.secret_key", ""),
            "api_type": config.get("invoice.type", "vat_invoice"),
            "region": config.get("invoice.region", ""),
        }
        
        try:
            engine = engine_class(engine_config)
            return engine
        except Exception as e:
            logger.error(f"创建引擎失败: {e}", exc_info=True)
            return None
    
    def _format_invoice_result(
        self,
        result: Dict[str, Any],
        invoice_type: str
    ) -> Dict[str, Any]:
        """
        格式化发票识别结果
        
        将原始API返回结果格式化为统一的发票数据结构
        """
        if result.get("code") != 100:
            return result
        
        data = result.get("data", {})
        
        # 如果已经是格式化的数据，直接返回
        if result.get("type") == "invoice":
            return result
        
        # 获取发票类型信息
        invoice_info = INVOICE_TYPES.get(invoice_type, INVOICE_TYPES["invoice"])
        
        formatted = {
            "code": 100,
            "type": "invoice",
            "invoice_type": invoice_type,
            "invoice_name": invoice_info["name"],
            "data": data,
            "format": "json",
        }
        
        return formatted
    
    # ========================= 【qml接口】 =========================
    
    def getStatus(self) -> Dict[str, Any]:
        """返回当前状态"""
        return {
            "engineType": self._engine_type,
            "invoiceType": self._invoice_type,
            "resultsCount": len(self._results),
            "missionListsLength": self.getMissionListsLength(),
        }
    
    def setEngine(self, engine_type: str) -> str:
        """设置识别引擎"""
        if engine_type not in get_cloud_engines():
            return f"[Error] 不支持的引擎类型: {engine_type}"
        
        self._engine_type = engine_type
        self._engine = None  # 重置引擎，下次使用时重新创建
        return "[Success]"
    
    def setInvoiceType(self, invoice_type: str) -> str:
        """设置发票类型"""
        if invoice_type not in INVOICE_TYPES:
            return f"[Error] 不支持的发票类型: {invoice_type}"
        
        self._invoice_type = invoice_type
        return "[Success]"
    
    def getInvoiceTypes(self) -> Dict[str, Any]:
        """获取支持的发票类型"""
        return INVOICE_TYPES
    
    def getCloudEngines(self) -> List[Dict[str, str]]:
        """获取可用的云OCR引擎"""
        engines = []
        for engine_type in get_cloud_engines():
            engine_class = get_engine_class(engine_type)
            if engine_class:
                engines.append({
                    "value": engine_type,
                    "label": engine_class.ENGINE_NAME,
                })
        return engines
    
    def getResults(self) -> List[Dict[str, Any]]:
        """获取所有识别结果"""
        return self._results
    
    def clearResults(self):
        """清空识别结果"""
        self._results = []
    
    def exportResults(self, output_path: str, format_type: str = "json") -> Dict[str, Any]:
        """
        导出识别结果
        
        Args:
            output_path: 输出文件路径
            format_type: 输出格式 (json/csv/excel)
            
        Returns:
            导出结果
        """
        if not self._results:
            return {
                "code": 101,
                "data": "没有可导出的结果",
            }
        
        try:
            if format_type == "json":
                return self._export_json(output_path)
            elif format_type == "csv":
                return self._export_csv(output_path)
            elif format_type == "excel":
                return self._export_excel(output_path)
            else:
                return {
                    "code": 902,
                    "data": f"不支持的导出格式: {format_type}",
                }
        except Exception as e:
            logger.error(f"导出结果失败: {e}", exc_info=True)
            return {
                "code": 903,
                "data": f"[Error] 导出失败: {e}",
            }
    
    def _export_json(self, output_path: str) -> Dict[str, Any]:
        """导出为JSON"""
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(self._results, f, ensure_ascii=False, indent=2)
        
        return {
            "code": 100,
            "data": f"已导出 {len(self._results)} 条记录到: {output_path}",
        }
    
    def _export_csv(self, output_path: str) -> Dict[str, Any]:
        """导出为CSV"""
        import csv
        
        # 收集所有字段
        all_fields = set()
        for result in self._results:
            data = result.get("data", {})
            if isinstance(data, dict):
                all_fields.update(data.keys())
        
        all_fields = sorted(list(all_fields))
        
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            
            # 写入表头
            writer.writerow(["文件路径", "发票类型"] + all_fields)
            
            # 写入数据
            for result in self._results:
                row = [
                    result.get("path", ""),
                    result.get("invoice_name", ""),
                ]
                data = result.get("data", {})
                for field in all_fields:
                    value = data.get(field, "") if isinstance(data, dict) else ""
                    row.append(str(value))
                writer.writerow(row)
        
        return {
            "code": 100,
            "data": f"已导出 {len(self._results)} 条记录到: {output_path}",
        }
    
    def _export_excel(self, output_path: str) -> Dict[str, Any]:
        """导出为Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            return {
                "code": 904,
                "data": "[Error] 请先安装 openpyxl: pip install openpyxl",
            }
        
        wb = Workbook()
        ws = wb.active
        ws.title = "发票识别结果"
        
        # 收集所有字段
        all_fields = set()
        for result in self._results:
            data = result.get("data", {})
            if isinstance(data, dict):
                all_fields.update(data.keys())
        
        all_fields = sorted(list(all_fields))
        headers = ["文件路径", "发票类型"] + all_fields
        
        # 写入表头
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # 写入数据
        for row_idx, result in enumerate(self._results, 2):
            ws.cell(row=row_idx, column=1, value=result.get("path", ""))
            ws.cell(row=row_idx, column=2, value=result.get("invoice_name", ""))
            
            data = result.get("data", {})
            for col_idx, field in enumerate(all_fields, 3):
                value = data.get(field, "") if isinstance(data, dict) else ""
                ws.cell(row=row_idx, column=col_idx, value=str(value))
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(output_path)
        
        return {
            "code": 100,
            "data": f"已导出 {len(self._results)} 条记录到: {output_path}",
        }
    
    def recognizeSingle(
        self,
        image_path: str,
        invoice_type: str = None,
        engine_type: str = None,
        config: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """
        识别单张发票
        
        Args:
            image_path: 图片路径
            invoice_type: 发票类型（可选）
            engine_type: 引擎类型（可选）
            config: 引擎配置（可选）
            
        Returns:
            识别结果
        """
        # 使用指定参数或默认参数
        inv_type = invoice_type or self._invoice_type
        eng_type = engine_type or self._engine_type
        
        # 如果配置变化，重新创建引擎
        if config or eng_type != self._engine_type or not self._engine:
            cfg = config or {}
            cfg["invoice.type"] = inv_type
            self._engine_type = eng_type
            self._engine = self._create_engine(cfg)
            
            if not self._engine:
                return {
                    "code": 906,
                    "data": f"[Error] 引擎创建失败: {eng_type}",
                }
            
            self._engine.initialize()
        
        try:
            res = self._engine.recognize(image_path, api_type=inv_type)
            
            if res.get("code") == 100:
                res = self._format_invoice_result(res, inv_type)
                res["path"] = image_path
            
            return res
            
        except Exception as e:
            logger.error(f"单张发票识别失败: {e}", exc_info=True)
            return {
                "code": 907,
                "data": f"[Error] 识别失败: {e}",
            }


# 全局 发票识别任务管理器
MissionInvoice = __MissionInvoiceClass()
